import os.path

import PySimpleGUI as sg
import datetime
import re
from helper import consts
from layouts import laout_extender
from data_classes import document, database_class
from openpyxl import load_workbook


def validate_doc(document):
    current_year = datetime.datetime.now().year
    regex = re.compile('[@_!#$%^&*()"=+<>?/|,}{~:]')
    regex_char = re.compile('[@_!#$%^&*()"=+<>?/|,}{~:a-zA-Z]')
    doc_chars = ('G', 'F', 'g', 'f')
    doc_all_chars = ('G', 'F', 'g', 'f', '1', '2',
                     '3', '4', '5', '6', '7', '8', '9')
    is_valid = True
    value_invalid = []

    if regex_char.search(document[consts.key_doc_input_year]) is None:
        if document[consts.key_doc_input_year] == '' or (int(document[consts.key_doc_input_year]) < 1043) or (
                int(document[consts.key_doc_input_year]) > current_year):
            value_invalid.append('Year')
            is_valid = False

    else:
        value_invalid.append('Year')
        is_valid = False

    if len(document[consts.key_doc_input_doc]) < 2 and document[consts.key_doc_input_doc].startswith(doc_chars):
        value_invalid.append('Document no')
        is_valid = False

    if not document[consts.key_doc_input_doc].startswith(doc_all_chars):
        value_invalid.append('Document no')
        is_valid = False

    if regex.search(document[consts.key_doc_input_doc]) is not None:
        value_invalid.append('Document no')
        is_valid = False
    result = [is_valid, value_invalid]
    return result


def generate_doc_input_error_msg(values_invalid):
    error_msg = ''
    for item in values_invalid:
        error_msg = ('\n Invalid' + ':' + item)
    return error_msg


def save_document_no(value, list_of_doc, window, db_connection, hide_doc_lit, file=None):
    validation_result = validate_doc(document=value)

    if validation_result[0]:
        doc_num = value[consts.key_doc_input_doc] + '-' + value[consts.key_doc_input_book] + '-' + value[
            consts.key_doc_input_year]
        volume = value[consts.key_doc_input_vol]
        document_obj = document.Document(doc_num, )
        if document_obj not in list_of_doc:
            list_of_doc.append(document_obj)
            database_class.add_doc_to_primary_tb(db_connection, doc_no=doc_num,
                                                 volume=volume)  # add document to database
            if doc_num not in hide_doc_lit:
                window.extend_layout(window[consts.key_col_add_doc],
                                     laout_extender.add_doc_colum1(document_obj.get_doc_number(), vol=volume),
                                     )
            else:
                window[consts.key_doc_number_holder_frame_offset + doc_num].unhide_row()
                # window[consts.key_doc_number_holder_frame_offset + doc_num].unhide_row()
                window[consts.key_del_doc_frm_col_1 + doc_num].unhide_row()
            window[doc_num].set_cursor(cursor='hand2', cursor_color=None)
            window.refresh()
            window[consts.key_col_add_doc].contents_changed()

        else:
            sg.popup_ok(f'\nThe Document {doc_num} already exists\n', title='Stop', modal=True, keep_on_top=True,
                        font=consts.heading_fonts,
                        no_titlebar=True,
                        background_color='red')
    else:
        error_message = generate_doc_input_error_msg(validation_result[1])
        sg.popup(error_message, title='Error', modal=True, keep_on_top=True, font=consts.heading_fonts,
                 background_color='red')


def get_document_from_db(db_connection, window, ):
    thresh_dir = consts.thresh_dir
    raw_dir = consts.raw_dir
    procesed_dir = consts.processed_dir
    doc_db_list = database_class.get_doc_from_primary(db_connection)
    doc_list = []
    if doc_db_list:
        for doc in doc_db_list:
            # print(doc)
            raw_dir_path = os.path.join(raw_dir, doc[0])
            thresh_dir_path = os.path.join(thresh_dir, doc[0])
            procesed_dir_path = os.path.join(thresh_dir, doc[0])
            document_obj = document.Document(doc[0])  # set doc No
            document_obj.set_volume(doc[1])  # set volume
            if doc[2] == 1:
                document_obj.set_capture_started()
            if doc[3] == 1:
                document_obj.set_capture_completed(True)
            if doc[4] == 1:
                document_obj.set_process_started()
            if doc[5] == 1:
                document_obj.set_process_completed(True)
            if doc[6] == 1:
                document_obj.set_document_created(True)
            if doc[3] == 1:
                window.extend_layout(window[consts.key_col_add_doc],
                                     laout_extender.add_doc_colum1_with_cap_completed(document_obj.get_doc_number(),
                                                                                      document_obj.get_volume())
                                     )
            else:
                window.extend_layout(window[consts.key_col_add_doc],
                                     laout_extender.add_doc_colum1(document_obj.get_doc_number(),
                                                                   document_obj.get_volume())
                                     )
            window[document_obj.get_doc_number()].set_cursor(cursor='hand2', cursor_color=None)
            window.refresh()
            window[consts.key_col_add_doc].contents_changed()

            if doc[2] == 1 and (doc[3] == 0 or doc[3] == 1):  # capture start but not click the save all button
                if os.path.exists(thresh_dir_path):
                    for img in os.listdir(thresh_dir_path):
                        document_obj.add_raw_image(document.CapturedImages(os.path.join(thresh_dir_path, img)))
            # elif doc[3] == 1 and doc[4] == 0:
            #     dir_path = os.path.join(consts.raw_image_directory, doc[0])
            #     for img in os.listdir(dir_path):
            #         document_obj.add_raw_image(document.CapturedImages(os.path.join(dir_path, img)))

            elif doc[4] == 1 and doc[5] == 0:  # process start but not completed
                if os.path.exists(thresh_dir_path):
                    for img in os.listdir(thresh_dir_path):
                        document_obj.add_raw_image(document.CapturedImages(os.path.join(thresh_dir_path, img)))
                if os.path.exists(procesed_dir_path):
                    for img in os.listdir(procesed_dir_path):
                        document_obj.add_raw_image(document.CapturedImages(os.path.join(procesed_dir_path, img)))

            elif doc[5] == 1:
                if os.path.exists(procesed_dir_path):
                    for img in os.listdir(procesed_dir_path):
                        document_obj.add_raw_image(document.CapturedImages(os.path.join(procesed_dir_path, img)))

            try:
                document_obj.sort_raw_image_list()
            except:
                pass

            doc_list.append(document_obj)
    return doc_list


# collect information from excel file
def get_doc_from_file(file_name,list_of_doc, db_connection, hide_doc_lit, window):
    cclist = []
    workbook = load_workbook(file_name)
    sheet = workbook.active
    for row in sheet.iter_rows(values_only=True):
        cclist.append(row[2:6])
    for docs in cclist:
        if docs[0] is not None:
            doc_num = str(docs[0])+'-'+str(docs[1])+'-'+str(docs[2])
            document_obj = document.Document(doc_num)
            volume = docs[3]
            if document_obj not in list_of_doc:
                list_of_doc.append(document_obj)
                database_class.add_doc_to_primary_tb(db_connection, doc_no=doc_num,
                                                     volume=volume)  # add document to database
                if doc_num not in hide_doc_lit:
                    window.extend_layout(window[consts.key_col_add_doc],
                                         laout_extender.add_doc_colum1(document_obj.get_doc_number(), vol=volume),
                                         )
                else:
                    window[consts.key_doc_number_holder_frame_offset + doc_num].unhide_row()
                    # window[consts.key_doc_number_holder_frame_offset + doc_num].unhide_row()
                    window[consts.key_del_doc_frm_col_1 + doc_num].unhide_row()
                window[doc_num].set_cursor(cursor='hand2', cursor_color=None)
                window.refresh()
                window[consts.key_col_add_doc].contents_changed()
                # window[consts.key_col_add_doc].update()

                #
                # print(document_list)
            # else:
            #     sg.popup_ok(f'\nThe Document {doc_num} already exists\n', title='Stop', modal=True, keep_on_top=True,
            #                 font=consts.heading_fonts,
            #                 no_titlebar=True,
            #                 background_color='red')
            #     break


