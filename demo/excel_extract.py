# import pandas as pd
# data = pd.read_excel(r'E:\ccm\demo\cc_list.xlsx')
# df = pd.DataFrame(data, columns=['Docno','BookNo','Year','Volume'])
# print(df)

from openpyxl import load_workbook

workbook = load_workbook('E:\ccm\demo\cc_list.xlsx')

sheet = workbook.active
print(sheet["A2"].value)
print(sheet.title)
print(sheet.title)
cclist = []
for row in sheet.iter_rows(values_only=True):
    cclist.append(row[2:6])

print(cclist)